import asyncio
import logging
import pandas as pd

# from pprint import pprint

from openpyxl.workbook import Workbook
from openpyxl.styles import NamedStyle, Font, PatternFill, Border, Side, Alignment

from async_colab_module import (
    get_api_tokens,
    MoySklad,
    get_prime_cost,
    get_stock_for_bundle,
    get_ya_data_,
)
from async_colab_module.tabstyle import TabStyles
from async_colab_module.ya_market import (
    YM,
    get_ya_campaign_and_business_ids,
    chunked_offers_list,
    get_dict_for_commission,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("PRICES")


async def get_desired_prices(plan_margin: float = 25.0, fbs: bool = True):
    ms_token, _, ym_token = get_api_tokens()
    ms_client = MoySklad(api_key=ms_token)
    products_ = await ms_client.get_bundles()
    print(f"Мой склад: {len(products_)}")
    # Оставляем только Яндекс
    ms_ya_products = [
        product for product in products_ if "ЯндексМаркет" in product["pathName"]
    ]
    print("Мой склад: Получение остатка товара")
    stocks = await ms_client.get_stock()
    ms_stocks = {stock["assortmentId"]: stock["quantity"] for stock in stocks}
    print("Мой склад: Получение себестоимости товара")
    ms_ya_products_ = {
        product["article"]: {
            "STOCK": get_stock_for_bundle(ms_stocks, product),
            "PRIME_COST": get_prime_cost(product.get("salePrices", [])),
            "NAME": product["name"],
        }
        for product in ms_ya_products
    }
    logger.info(len(ms_ya_products_))
    await ms_client.close()

    ym_client = YM(api_key=ym_token, max_rete=45, time_period=3)

    campaign_id, business_id = await get_ya_campaign_and_business_ids(
        ym_client, fbs=fbs
    )

    offers = await ym_client.get_full_offers(business_id)

    print("ЯндексМаркет: Получение актуальных тарифов")
    offers_commission_dict = await chunked_offers_list(
        get_dict_for_commission,
        ym_client=ym_client,
        campaign_id=campaign_id,
        data=offers,
        chunk_size=200,
    )

    await ym_client.close()

    ya_set = set(offers_commission_dict)
    ms_set = set(ms_ya_products_)

    result_dict = {
        key: {**offers_commission_dict.get(key, {}), **ms_ya_products_.get(key, {})}
        for key in ya_set & ms_set
    }

    ya_ms_set = ya_set - ms_set
    if ya_ms_set:
        print("Номенклатура которая есть в ЯндексМаркете, но не связана в МС:")
        print("\n".join(ya_ms_set))

    data_for_report = [
        get_ya_data_(article, result_dict[article], plan_margin)
        for article in result_dict
    ]
    print('Формирую отчет "Рекомендуемые цены"')
    # progress_bar.update(50)
    pd.set_option("display.max_columns", None)
    pd.set_option("display.max_rows", None)
    df = pd.DataFrame(data_for_report)
    df_total = (
        df.agg(
            {
                "stock": "sum",
                "price": "sum",
                "recommended_price": "sum",
                "prime_cost": "sum",
                "commission": "sum",
                "acquiring": "sum",
                "delivery": "sum",
                "crossregional_delivery": "sum",
                "sorting": "sum",
                "profit": "sum",
            }
        )
        .to_frame()
        .T
    )
    df_total["name"] = ""
    df_total["article"] = ""
    df_total["profitability"] = round((df_total["profit"] / df_total["price"]) * 100, 1)

    df = pd.concat([df, df_total], ignore_index=True)

    df.columns = [
        "Номенклатура",
        "Артикул",
        "Остаток",
        "Текущая цена",
        "Рекомендуемая цена",
        "Себестоимость",
        "Комиссия",
        "Эквайринг",
        "Доставка",
        "Доставка в округ",
        "Обработка",
        "Прибыль",
        "Рентабельность",
    ]
    # print(df)
    path_xls_file = f'ya_{"fbs" if fbs else "express"}_рекомендуемые_цены.xlsx'
    style = ExcelStyle()
    style.style_dataframe(df, path_xls_file, "Номенклатура YA")
    print("Файл отчета готов")
    files.download(path_xls_file)


class ExcelStyle:
    def __init__(
        self,
        header_font=None,
        header_fill=None,
        header_border=None,
        cell_font=None,
        cell_fill=None,
        cell_border=None,
    ):

        self.border_side = Side(border_style="thin", color="C5B775")

        self.header_style = NamedStyle(name="header_style")
        self.cell_style = NamedStyle(name="cell_style")

        self.header_style.font = (
            header_font if header_font else Font(name="Calibri", bold=True)
        )
        self.header_style.fill = (
            header_fill if header_fill else PatternFill("solid", fgColor="F4ECC5")
        )
        self.header_style.border = (
            header_border
            if header_border
            else Border(
                left=self.border_side,
                right=self.border_side,
                top=self.border_side,
                bottom=self.border_side,
            )
        )

        self.cell_style.font = cell_font if cell_font else Font(name="Calibri")
        self.cell_style.fill = cell_fill if cell_fill else PatternFill()
        self.cell_style.border = (
            cell_border
            if cell_border
            else Border(
                left=self.border_side,
                right=self.border_side,
                top=self.border_side,
                bottom=self.border_side,
            )
        )

    def apply_to_workbook(self, workbook: Workbook):
        if "header_style" not in workbook.style_names:
            workbook.add_named_style(self.header_style)
        if "cell_style" not in workbook.style_names:
            workbook.add_named_style(self.cell_style)

    def style_dataframe(self, df: pd.DataFrame, file_path: str, sheet_title: str):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sheet_title
        sheet.sheet_format.defaultColWidth = 15
        # Установите ширину для определенных столбцов
        sheet.column_dimensions["A"].width = 30
        sheet.column_dimensions["B"].width = 18
        sheet.column_dimensions["C"].width = 10
        tab_styles = TabStyles()

        self.apply_to_workbook(workbook)

        columns_to_align_right = [7, 8, 9, 10, 11, 12, 13]

        for col_idx, column in enumerate(df.columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=column)
            cell.style = self.header_style

        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.style = self.cell_style

                if col_idx in columns_to_align_right:
                    cell.alignment = tab_styles.columns_to_align_right
                    cell.number_format = (
                        "#,##0.0"  # Формат с одним знаком после запятой
                    )

        sheet.auto_filter.ref = sheet.dimensions
        workbook.save(file_path)


if __name__ == "__main__":

    async def main():
        await get_desired_prices()

    asyncio.run(main())
